"""Turning an uploaded file into text the model can read.

Ask Alaiy's model is reached through `engine.llm.complete()`, which today fronts
a Gemini deployment behind an Anthropic-compatible gateway. Native `image` and
`document` content blocks are not reliably supported across that seam, so an
attachment is never passed to the model as a file. It reaches the conversation
one of two ways, chosen in `runner._attachment_blocks`:

  - As a *pointer*, when the user can call FAC's `extract_file_content` — the
    model reads the file itself, with no ceiling on how much of it it can
    reach, and gets that tool's table extraction and OCR into the bargain.
    This is the better path and the one that runs wherever FAC is available.
  - As *inlined text*, extracted here and quoted into the user's message as an
    ordinary `text` block. The fallback, and why this module exists: FAC is
    deliberately optional (see `hooks.py`), so a site without it — or with the
    data_science plugin disabled — still gets working attachments.

Either way history replay, trimming and the tool loop need no notion of a file.

The extraction below therefore duplicates a little of what `extract_file_content`
does, on purpose. It is the floor, not the ceiling: fewer formats, no OCR, no
table structure, capped at MAX_CHARS_PER_FILE. It also runs at upload time on
both paths, because it is what validates that a file is readable at all before
the user sends it.

Everything in this module is pure — bytes in, text out, no Frappe documents — so
it can be unit-tested without a site. The Frappe side lives in `api/chat.py`
(upload) and `runner.py` (inlining).

Extraction runs inside the *upload* web request, not the send path:

  - `send_message` only enqueues a turn and must stay fast; a 10 MB PDF parsed
    there would block the response for seconds.
  - "This PDF is scanned, there is no text in it" has to reach the user while
    they can still remove the file, i.e. before they hit send.

That places a parser on a web worker, which is why every limit below is a hard
cap rather than a suggestion: an `.xlsx` is a zip archive, and a decompression
bomb is trivial to build.
"""

import csv
import io
import os

import frappe

# A file the user uploads. Frappe's own `conf.max_file_size` defaults to the
# same figure; checking ours first means the user gets our message, not a
# MaxFileSizeReachedError. See `upload_limit` — a site that lowers its own limit
# lowers this one with it, or the better message would be the one that never
# fires.
MAX_UPLOAD_BYTES = 10 * 1024 * 1024

# Extracted text per file, and summed across one message. The per-file figure
# matches `runner.MAX_TOOL_RESULT_CHARS` deliberately: both answer the same
# question — how much of one artefact is worth paying to put in the context.
MAX_CHARS_PER_FILE = 20_000
MAX_CHARS_PER_MESSAGE = 40_000

MAX_ATTACHMENTS_PER_MESSAGE = 5

# Structural caps, applied before the char cap can be reached. These bound the
# *work*, not just the output: a 50,000-page PDF would otherwise be parsed in
# full only to have all but the first few pages thrown away.
MAX_PDF_PAGES = 200
MAX_SHEET_ROWS = 2_000

TRUNCATION_MARKER = "\n… [truncated, {total} chars total]"

# Extension is the allowlist, not the browser-supplied content type — that is
# attacker-controlled. No image formats: see the module docstring.
PLAIN_EXTENSIONS = {
	".txt",
	".md",
	".json",
	".yaml",
	".yml",
	".py",
	".js",
	".ts",
	".sql",
	".log",
	".xml",
	".html",
	".htm",
	".css",
	".ini",
	".cfg",
	".toml",
}
SHEET_EXTENSIONS = {".xlsx", ".xlsm"}
DELIMITED_EXTENSIONS = {".csv": ",", ".tsv": "\t"}

SUPPORTED = PLAIN_EXTENSIONS | SHEET_EXTENSIONS | set(DELIMITED_EXTENSIONS) | {".pdf"}


def extension_of(file_name):
	return os.path.splitext(file_name or "")[1].lower()


def check_supported(file_name):
	"""Validate the extension, returning it. Throws with a usable message."""
	ext = extension_of(file_name)

	if ext == ".xls":
		# openpyxl reads the OOXML format only; the legacy binary .xls needs a
		# separate library we do not carry. Say so rather than "unsupported".
		frappe.throw(
			frappe._("{0} is in the old Excel format. Save it as .xlsx and attach it again.").format(
				file_name
			)
		)

	if ext not in SUPPORTED:
		frappe.throw(
			frappe._("{0} can't be read. Supported files: {1}.").format(
				file_name or frappe._("This file"), ", ".join(sorted(SUPPORTED))
			)
		)

	return ext


def upload_limit():
	"""The effective size ceiling: ours, or the site's, whichever is lower.

	`save_file` enforces `conf.max_file_size` itself and throws its own generic
	error. The only thing checking size here buys is a message that names the
	file and the real limit — which a site that sets a *lower* max_file_size
	would silently lose, since Frappe's throw would fire first on everything
	between the two figures.
	"""
	from frappe.utils.file_manager import get_max_file_size

	return min(MAX_UPLOAD_BYTES, get_max_file_size())


def check_size(file_name, size):
	limit = upload_limit()
	if size > limit:
		frappe.throw(
			frappe._("{0} is {1} MB. The limit is {2} MB.").format(
				file_name, round(size / 1024 / 1024, 1), round(limit / 1024 / 1024, 1)
			)
		)
	if not size:
		frappe.throw(frappe._("{0} is empty.").format(file_name))


def extract(file_name, content):
	"""Text of one uploaded file, capped. `content` is bytes.

	Never hand this a `File.get_content()` result. That method walks a list of
	encodings and returns the first that does not raise; windows-1252 decodes
	almost any byte sequence, so a PDF comes back as mojibake `str` rather than
	failing. Read the bytes yourself.
	"""
	ext = check_supported(file_name)

	if ext == ".pdf":
		text = _pdf(content)
	elif ext in SHEET_EXTENSIONS:
		text = _xlsx(content)
	elif ext in DELIMITED_EXTENSIONS:
		text = _delimited(content, DELIMITED_EXTENSIONS[ext])
	else:
		text = _plain(content)

	text = (text or "").strip()
	if not text:
		# Overwhelmingly this is a scanned PDF — pages of images with no text
		# layer. Sending an empty attachment would have the model confidently
		# answer about a document it cannot see, so fail here instead.
		frappe.throw(
			frappe._("No text could be read from {0}. If it is a scan, it needs OCR first.").format(
				file_name
			)
		)

	return _cap(text, MAX_CHARS_PER_FILE)


# ── Per-format readers ───────────────────────────────────────────────────────
def _pdf(content):
	from pypdf import PdfReader

	try:
		reader = PdfReader(io.BytesIO(content))
	except Exception as exc:
		frappe.throw(frappe._("This PDF could not be opened: {0}").format(exc))

	if getattr(reader, "is_encrypted", False):
		# An empty-password PDF decrypts silently; a real one cannot be read.
		try:
			reader.decrypt("")
		except Exception:
			frappe.throw(frappe._("This PDF is password-protected."))

	parts = []
	total = 0
	for index, page in enumerate(reader.pages):
		if index >= MAX_PDF_PAGES or total >= MAX_CHARS_PER_FILE:
			break
		try:
			page_text = (page.extract_text() or "").strip()
		except Exception:
			# One malformed page should not lose the other 199.
			continue
		if not page_text:
			continue
		parts.append(f"--- page {index + 1} ---\n{page_text}")
		total += len(page_text)

	return "\n\n".join(parts)


def _xlsx(content):
	from openpyxl import load_workbook

	try:
		# read_only streams rows instead of building the whole object graph;
		# data_only gives us cached formula *values*, since the formula text is
		# no use to the model.
		book = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
	except Exception as exc:
		frappe.throw(frappe._("This spreadsheet could not be opened: {0}").format(exc))

	parts = []
	total = 0
	try:
		for sheet in book.worksheets:
			if total >= MAX_CHARS_PER_FILE:
				break
			rows = []
			for index, row in enumerate(sheet.iter_rows(values_only=True)):
				if index >= MAX_SHEET_ROWS:
					rows.append(f"… [{sheet.title} truncated at {MAX_SHEET_ROWS} rows]")
					break
				line = "\t".join("" if cell is None else str(cell) for cell in row).rstrip("\t")
				if line:
					rows.append(line)
			if not rows:
				continue
			block = f"--- sheet: {sheet.title} ---\n" + "\n".join(rows)
			parts.append(block)
			total += len(block)
	finally:
		book.close()

	return "\n\n".join(parts)


def _delimited(content, delimiter):
	text = _plain(content)
	rows = []
	for index, row in enumerate(csv.reader(io.StringIO(text), delimiter=delimiter)):
		if index >= MAX_SHEET_ROWS:
			rows.append(f"… [truncated at {MAX_SHEET_ROWS} rows]")
			break
		rows.append("\t".join(row))
	return "\n".join(rows)


def _plain(content):
	"""Decode text bytes, guessing the encoding when it is not UTF-8."""
	try:
		return content.decode("utf-8")
	except UnicodeDecodeError:
		pass

	import chardet

	# Sniffing the whole of a 10 MB file is slow and no more accurate than a
	# generous prefix.
	guess = chardet.detect(content[:200_000]) or {}
	encoding = guess.get("encoding") or "utf-8"
	return content.decode(encoding, errors="replace")


def _cap(text, limit):
	if len(text) <= limit:
		return text
	return text[:limit] + TRUNCATION_MARKER.format(total=len(text))


# ── Blocks ───────────────────────────────────────────────────────────────────
# FAC's file reader, in the data_science plugin. When the user can call it, an
# attachment is passed as a *reference* and the model reads what it needs, which
# beats inlining a truncated prefix: the tool does table extraction and OCR, and
# nothing is capped at MAX_CHARS_PER_FILE. See `runner._attachment_blocks`.
FILE_TOOL = "extract_file_content"

# Its companion in the same plugin — a sandboxed Python interpreter. Only worth
# mentioning to the model for tabular files, and only when it is actually there.
CODE_TOOL = "run_python_code"


def render_block(file_name, text):
	"""The content block that carries one attachment into the conversation.

	Delimited rather than merged into the user's prose so the model can tell the
	document from the question, and can cite the file by name.
	"""
	return {
		"type": "text",
		"text": f'<attachment name="{file_name}">\n{text}\n</attachment>',
	}


def pointer_block(meta, can_run_code=False):
	"""An attachment as a reference for the model to go and read.

	Costs a few dozen tokens instead of up to 20k, and — unlike an inlined
	prefix — puts no ceiling on how much of the file the model can reach. It
	stays in the history verbatim rather than being elided, because the file_url
	is the only way back to the file on a later turn.
	"""
	ext = extension_of(meta.get("file_name"))
	if ext in SHEET_EXTENSIONS or ext in DELIMITED_EXTENSIONS:
		how = "Use operation \"parse_data\" for the rows"
		if can_run_code:
			how += f', then {CODE_TOOL} if you need to compute over them rather than read them'
	elif ext == ".pdf":
		how = 'Use operation "extract" for the text, or "extract_tables" if the answer is in a table'
	else:
		how = 'Use operation "extract"'

	return {
		"type": "text",
		"text": (
			f'<attachment name="{meta.get("file_name")}" file_url="{meta.get("file_url")}"/>\n'
			f"The user attached this file. Its contents are not included here — read it with "
			f"the {FILE_TOOL} tool, passing that exact file_url. {how}. "
			f"Do not answer questions about this file without reading it first."
		),
	}


def stub_block(meta):
	"""What an attachment shrinks to once it is no longer the current one.

	See `runner._elide_old_attachments`. It names the file so the model can ask
	for it back rather than inventing its contents.
	"""
	return {
		"type": "text",
		"text": (
			f'<attachment name="{meta.get("file_name")}" chars="{meta.get("chars") or 0}" dropped="true"/>\n'
			"The contents of this earlier attachment are no longer in context. "
			"Ask the user to attach it again if you need it."
		),
	}
