"""Turning a table the model has already read into a file the user can download.

The mirror image of `attachments.py`. That module is bytes → text, inbound; this
one is rows → bytes, outbound. Both are pure — a normalised table in, bytes out,
no Frappe documents — so both can be exercised without a site. The Frappe side of
this one (the `File` row, the permission, the chip) lives in `artifacts.py`.

## Where the rows come from

The model passes them, as tool arguments. It does *not* name an earlier tool call
for the server to re-run, and that is the load-bearing decision here:

  - There is no handle to name. Tool results are not stored as retrievable
    values, and `runner._run_tools` truncates them at `MAX_TOOL_RESULT_CHARS`, so
    the model would be exporting a result it only ever saw a prefix of.
  - Re-running is a second execution of a live tool: a second audit-log row, a
    second permission evaluation, and a different answer if the data moved in
    between. The file would then disagree with the reply above it.
  - Inline is honest. The file contains exactly the numbers in the message, which
    is the only property that makes a download trustworthy in a system whose
    system prompt already says "never invent figures".

What it gives up is a genuinely bulk export — "every order last month" — which is
refused rather than served. That is the right thing to give up first: the answer
for fifty thousand rows is a scheduled report, not a chat message. The throw says
so, and names the format to use instead where one would help.

## Why the caps are shaped the way they are

The *first* cap on any of this is the model's own output limit, since every cell
costs output tokens to write. Everything below exists for the model that tries
anyway, so it gets a sentence back instead of pinning a worker.
"""

import csv
import io
import os
import re

import frappe

FORMATS = ("xlsx", "csv", "pdf")

# Cells is the real bound — rows and columns are the legible ones, and a caller
# reads them first.
MAX_ROWS = 5_000
MAX_COLUMNS = 50
MAX_CELLS = 50_000

# Truncated in place rather than thrown on: a long description cell is a shape of
# data, not an abuse of the tool.
MAX_CELL_CHARS = 500

# Much lower than the rest, and the only cap bounding a *subprocess*: `get_pdf`
# shells out to wkhtmltopdf with no timeout parameter, so page count is all that
# stands between a wide table and a pinned worker.
MAX_PDF_ROWS = 500

# Checked after generation, which is the only point at which a pdf's real cost is
# knowable — a 400-row table can compress to 40 KB or expand past a megabyte
# depending on how much text is in it.
MAX_BYTES = 5 * 1024 * 1024

# A model-suggested file name is a suggestion, not a path.
MAX_FILE_NAME_CHARS = 80

EXTENSIONS = {"xlsx": ".xlsx", "csv": ".csv", "pdf": ".pdf"}

# Anything that is not plainly part of a file name. Separators go too: this
# string reaches `save_file`, and "../" in it is a path, not a name.
_UNSAFE_NAME = re.compile(r"[^A-Za-z0-9 ._+-]+")


def check_format(fmt):
	"""Validate the format, returning it normalised. Throws with a usable message."""
	value = (fmt or "").strip().lower().lstrip(".")
	if value == "xls":
		# openpyxl writes OOXML only, and nothing downstream would open the
		# legacy binary format anyway. Say which one to ask for.
		frappe.throw(frappe._('Use format "xlsx" — the old .xls format is not written.'))
	if value not in FORMATS:
		frappe.throw(
			frappe._("{0} is not a format I can write. Use one of: {1}.").format(
				fmt or frappe._("That"), ", ".join(FORMATS)
			)
		)
	return value


def normalise(columns, rows, fmt):
	"""`(columns, rows)` as strings, every cap applied. Throws when one is hit.

	One validation for all three writers, called before any of them, so a caller
	cannot reach a writer with a table it should have been refused. Every throw
	names the real number and — where switching would help — the format to switch
	to, because a `tool_result` marked `is_error` is the model's only route to
	recovering from this.
	"""
	columns = [_cell(value) for value in (columns or [])]
	if not columns:
		frappe.throw(frappe._("A file needs at least one column. Pass the column headings."))
	if len(columns) > MAX_COLUMNS:
		frappe.throw(
			frappe._("{0} columns is more than the {1} a file can hold. Drop some.").format(
				len(columns), MAX_COLUMNS
			)
		)

	table = []
	for row in rows or []:
		if not isinstance(row, (list, tuple)):
			frappe.throw(frappe._("Every row must be a list of values, one per column."))
		# Short rows are padded and long ones cut: a model that loses a trailing
		# empty cell should get its file, not an error about arithmetic.
		values = [_cell(value) for value in row[: len(columns)]]
		values += [""] * (len(columns) - len(values))
		table.append(values)

	if not table:
		# Note what this does NOT say. "Retrieve the data first" reads as an
		# instruction to go and look again, which is how a model that has already
		# correctly established there is nothing to report ends up burning every
		# remaining turn hunting for a second source. An empty result is an
		# answer; the model needs telling to report it, not to retry.
		frappe.throw(
			frappe._(
				"There are no rows, so there is nothing to put in a file. Tell the user "
				"the result was empty and what you searched — do not look for another "
				"way to fill it."
			)
		)

	limit = MAX_PDF_ROWS if fmt == "pdf" else MAX_ROWS
	if len(table) > limit:
		if fmt == "pdf":
			frappe.throw(
				frappe._(
					"{0} rows is too many for a pdf — that is a printable document, not a "
					'dataset. The limit is {1}. Use format "xlsx" for this, or narrow the range.'
				).format(len(table), limit)
			)
		frappe.throw(
			frappe._(
				"{0} rows is more than the {1} one file can hold. Narrow the range, or "
				"summarise and export the summary."
			).format(len(table), limit)
		)

	cells = len(table) * len(columns)
	if cells > MAX_CELLS:
		frappe.throw(
			frappe._(
				"{0} rows × {1} columns is {2} cells; the limit is {3}. Drop some columns "
				"or narrow the range."
			).format(len(table), len(columns), cells, MAX_CELLS)
		)

	return columns, table


def safe_file_name(name, fmt):
	"""A file name from a language model, made safe and given the right extension."""
	base = os.path.basename(str(name or "").strip()) or "download"
	base = os.path.splitext(base)[0]
	base = _UNSAFE_NAME.sub("-", base).strip(" .-") or "download"
	return base[:MAX_FILE_NAME_CHARS] + EXTENSIONS[fmt]


def write(fmt, columns, rows, title=None, generated=None):
	"""Bytes of one file. `columns`/`rows` must have been through `normalise`."""
	if fmt == "xlsx":
		return write_xlsx(columns, rows, title)
	if fmt == "csv":
		return write_csv(columns, rows)
	return write_pdf(columns, rows, title, generated=generated)


def write_xlsx(columns, rows, title=None):
	from openpyxl import Workbook
	from openpyxl.styles import Font
	from openpyxl.utils import get_column_letter

	# write_only streams rows out instead of building the whole object graph —
	# the same reasoning as `read_only` on the way in. The catch is that a
	# write-only sheet serialises its header the moment the first row is
	# appended, so `freeze_panes` and `column_dimensions` have to be set BEFORE
	# that or they are silently dropped. Hence the width scan up front rather
	# than as we go: the rows are already fully in memory (they arrived as tool
	# arguments), so there is nothing to stream them out of anyway.
	book = Workbook(write_only=True)
	sheet = book.create_sheet(_sheet_title(title))

	widths = [len(heading) for heading in columns]
	for row in rows:
		for index, value in enumerate(row):
			if len(value) > widths[index]:
				widths[index] = len(value)
	for index, width in enumerate(widths):
		sheet.column_dimensions[get_column_letter(index + 1)].width = min(max(width + 2, 8), 60)
	# Keeps the headings visible while scrolling a long export, which is the
	# whole reason anyone asked for xlsx rather than csv.
	sheet.freeze_panes = "A2"

	bold = Font(bold=True)
	header = []
	for heading in columns:
		cell = _write_only_cell(sheet, heading)
		cell.font = bold
		header.append(cell)
	sheet.append(header)

	for row in rows:
		sheet.append([_typed(value) for value in row])

	buffer = io.BytesIO()
	book.save(buffer)
	return buffer.getvalue()


def write_csv(columns, rows):
	buffer = io.StringIO()
	writer = csv.writer(buffer)
	writer.writerow(columns)
	writer.writerows(rows)
	# utf-8-sig, deliberately. These files are opened in Excel on a Windows
	# desktop, and without the BOM every non-ASCII brand name is mojibake in the
	# one place the file actually gets read.
	return buffer.getvalue().encode("utf-8-sig")


def write_pdf(columns, rows, title=None, generated=None):
	from frappe.utils.pdf import get_pdf

	# Landscape past six columns: a wide table clipped at the margin is a broken
	# artefact, not a narrow one.
	orientation = "Landscape" if len(columns) > 6 else "Portrait"
	return get_pdf(
		_pdf_html(columns, rows, title, generated),
		{
			"orientation": orientation,
			# Passed explicitly for two reasons. `prepare_options` otherwise reads
			# `Print Settings.pdf_page_size` from the database on every single
			# export — and that setting exists for a site's invoices and letter
			# heads, which have nothing to do with the page a data table wants.
			"page-size": "A4",
		},
	)


def _pdf_html(columns, rows, title=None, generated=None):
	"""One self-contained document. No external assets, by necessity.

	`generated` is passed in rather than read from the clock: the site's timezone
	lives behind `frappe.get_system_settings`, and reaching for it here would
	make this module need a site — which is the one property it is trading
	everything else for.

	`get_pdf` passes `--disable-local-file-access` to wkhtmltopdf, so a
	stylesheet or image reference would be dropped without an error rather than
	fetched. Everything is inline, and every value is escaped: these strings came
	from a language model.
	"""
	from frappe.utils import escape_html

	heading = escape_html(title or "Export")
	stamp = f" · generated {escape_html(generated)}" if generated else ""
	head = "".join(f"<th>{escape_html(value)}</th>" for value in columns)
	body = "".join(
		"<tr>" + "".join(f"<td>{escape_html(value)}</td>" for value in row) + "</tr>"
		for row in rows
	)
	return (
		"<!doctype html><html><head><meta charset='utf-8'><style>"
		"body{font-family:Helvetica,Arial,sans-serif;font-size:9pt;color:#111;margin:0}"
		"h1{font-size:13pt;margin:0 0 2mm}"
		".meta{font-size:8pt;color:#666;margin:0 0 4mm}"
		"table{width:100%;border-collapse:collapse}"
		"th,td{border:0.4pt solid #bbb;padding:1.4mm 2mm;text-align:left;vertical-align:top}"
		"th{background:#f2f2f2;font-weight:bold}"
		"thead{display:table-header-group}"  # repeats the header on every page
		"tr{page-break-inside:avoid}"
		"</style></head><body>"
		f"<h1>{heading}</h1>"
		f"<p class='meta'>{len(rows)} rows{stamp}</p>"
		f"<table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table>"
		"</body></html>"
	)


# ── Cells ────────────────────────────────────────────────────────────────────
def _cell(value):
	"""One value as a capped string. `None` is empty, not "None"."""
	if value is None:
		return ""
	if isinstance(value, bool):
		# str(True) is "True", which is right, but only by accident — pin it.
		return "Yes" if value else "No"
	text = value if isinstance(value, str) else str(value)
	if len(text) > MAX_CELL_CHARS:
		return text[:MAX_CELL_CHARS] + "…"
	return text


def _typed(value):
	"""A cell as a number where it plainly is one, so a spreadsheet can sum it.

	Only for xlsx. Everything arrives as a string because `normalise` stringifies
	for one comparable definition of "cell", but writing "1234.5" as text turns
	every figure in the file into something the user has to convert by hand — the
	single most annoying property a generated spreadsheet can have.
	"""
	if not value or len(value) > 32:
		return value
	# A leading zero is an identifier, not a quantity. "007" and "0012345" are
	# SKUs and order numbers, and a spreadsheet that silently renders them as 7
	# and 12345 has destroyed the one column the user needs to look things up by.
	if value[0] == "0" and value not in ("0",) and not value.startswith("0."):
		return value
	try:
		# int first: `float("7")` would make every count a decimal.
		return int(value)
	except ValueError:
		pass
	try:
		number = float(value)
	except ValueError:
		return value
	# inf and nan round-trip through float() and then break the writer.
	return number if number == number and abs(number) != float("inf") else value


def _write_only_cell(sheet, value):
	from openpyxl.cell import WriteOnlyCell

	return WriteOnlyCell(sheet, value=value)


def _sheet_title(title):
	"""Excel rejects several characters in a sheet name and caps it at 31 chars."""
	name = re.sub(r"[\[\]:*?/\\]", " ", str(title or "Export")).strip()
	return (name or "Export")[:31]
